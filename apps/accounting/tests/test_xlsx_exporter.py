"""TDD tests for XLSX Export Adapter.

These tests verify the adapter produces valid, well-structured Excel workbooks
from approved invoice records without any real file I/O.

Run with:
    pytest tests/test_xlsx_exporter.py -v
"""

from __future__ import annotations

import io
import sys
import unittest
from pathlib import Path


import openpyxl

from platform_adapters.exporters.xlsx_exporter import XLSXExportAdapter
from platform_core.domain import ExportArtifact


# ---------------------------------------------------------------------------
# Sample records (approved invoice data as plain dicts)
# ---------------------------------------------------------------------------

INVOICE_1 = {
    "source_file_id": "f-001",
    "supplier_name": "ACME Corp",
    "supplier_tax_id": "0123456789",
    "invoice_number": "INV-001",
    "invoice_date": "2026-08-01",
    "currency": "USD",
    "subtotal": 900.0,
    "tax_amount": 90.0,
    "total_amount": 990.0,
    "items": [
        {"description": "Widget A", "quantity": 2, "unit_price": 300.0, "amount": 600.0},
        {"description": "Widget B", "quantity": 3, "unit_price": 100.0, "amount": 300.0},
    ],
}

INVOICE_2 = {
    "source_file_id": "f-002",
    "supplier_name": "Beta Ltd",
    "supplier_tax_id": None,
    "invoice_number": "INV-002",
    "invoice_date": "2026-08-02",
    "currency": "VND",
    "subtotal": None,
    "tax_amount": None,
    "total_amount": 1_500_000.0,
    "items": [
        {"description": "Service Fee", "quantity": 1, "unit_price": 1_500_000.0, "amount": 1_500_000.0},
    ],
}


def _load_workbook_from_artifact(artifact: ExportArtifact) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(artifact.content))


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestXLSXExportAdapterOutput(unittest.TestCase):

    def setUp(self):
        self.adapter = XLSXExportAdapter()

    def test_returns_export_artifact_type(self):
        result = self.adapter.export([INVOICE_1])
        self.assertIsInstance(result, ExportArtifact)

    def test_artifact_media_type_is_xlsx(self):
        result = self.adapter.export([INVOICE_1])
        self.assertEqual(result.media_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def test_artifact_content_is_non_empty_bytes(self):
        result = self.adapter.export([INVOICE_1])
        self.assertIsInstance(result.content, bytes)
        self.assertGreater(len(result.content), 0)

    def test_artifact_name_ends_with_xlsx(self):
        result = self.adapter.export([INVOICE_1])
        self.assertTrue(result.name.endswith(".xlsx"))

    def test_workbook_has_invoices_sheet(self):
        result = self.adapter.export([INVOICE_1])
        wb = _load_workbook_from_artifact(result)
        self.assertIn("Invoices", wb.sheetnames)

    def test_workbook_has_line_items_sheet(self):
        result = self.adapter.export([INVOICE_1])
        wb = _load_workbook_from_artifact(result)
        self.assertIn("Line Items", wb.sheetnames)


class TestXLSXInvoicesSheet(unittest.TestCase):

    def setUp(self):
        self.adapter = XLSXExportAdapter()
        self.artifact = self.adapter.export([INVOICE_1, INVOICE_2])
        self.wb = _load_workbook_from_artifact(self.artifact)
        self.ws = self.wb["Invoices"]

    def test_header_row_contains_supplier_name(self):
        headers = [cell.value for cell in self.ws[1]]
        self.assertIn("Supplier Name", headers)

    def test_header_row_contains_total_amount(self):
        headers = [cell.value for cell in self.ws[1]]
        self.assertIn("Total Amount", headers)

    def test_header_row_contains_currency(self):
        headers = [cell.value for cell in self.ws[1]]
        self.assertIn("Currency", headers)

    def test_data_rows_match_invoice_count(self):
        # Row 1 = header, rows 2+ = data
        data_rows = [r for r in self.ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)]
        self.assertEqual(len(data_rows), 2)

    def test_first_invoice_supplier_name_is_correct(self):
        headers = [cell.value for cell in self.ws[1]]
        col_idx = headers.index("Supplier Name")
        cell_value = self.ws.cell(row=2, column=col_idx + 1).value
        self.assertEqual(cell_value, "ACME Corp")

    def test_total_amount_is_numeric(self):
        headers = [cell.value for cell in self.ws[1]]
        col_idx = headers.index("Total Amount")
        cell_value = self.ws.cell(row=2, column=col_idx + 1).value
        self.assertIsInstance(cell_value, (int, float))


class TestXLSXLineItemsSheet(unittest.TestCase):

    def setUp(self):
        self.adapter = XLSXExportAdapter()
        self.artifact = self.adapter.export([INVOICE_1, INVOICE_2])
        self.wb = _load_workbook_from_artifact(self.artifact)
        self.ws = self.wb["Line Items"]

    def test_header_row_contains_description(self):
        headers = [cell.value for cell in self.ws[1]]
        self.assertIn("Description", headers)

    def test_header_row_contains_amount(self):
        headers = [cell.value for cell in self.ws[1]]
        self.assertIn("Amount", headers)

    def test_total_item_rows_matches_all_items(self):
        # INVOICE_1 has 2 items, INVOICE_2 has 1 item → 3 total
        data_rows = [r for r in self.ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)]
        self.assertEqual(len(data_rows), 3)

    def test_item_description_is_correct(self):
        headers = [cell.value for cell in self.ws[1]]
        col_idx = headers.index("Description")
        cell_value = self.ws.cell(row=2, column=col_idx + 1).value
        self.assertEqual(cell_value, "Widget A")


class TestXLSXEdgeCases(unittest.TestCase):

    def setUp(self):
        self.adapter = XLSXExportAdapter()

    def test_empty_records_produces_valid_xlsx(self):
        result = self.adapter.export([])
        wb = _load_workbook_from_artifact(result)
        self.assertIn("Invoices", wb.sheetnames)

    def test_invoice_with_no_items_is_handled(self):
        record = {**INVOICE_1, "items": []}
        result = self.adapter.export([record])
        wb = _load_workbook_from_artifact(result)
        ws = wb["Line Items"]
        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)]
        self.assertEqual(len(data_rows), 0)

    def test_none_fields_do_not_crash(self):
        record = {
            "source_file_id": "f-x",
            "supplier_name": None,
            "supplier_tax_id": None,
            "invoice_number": None,
            "invoice_date": None,
            "currency": None,
            "subtotal": None,
            "tax_amount": None,
            "total_amount": None,
            "items": [],
        }
        result = self.adapter.export([record])
        self.assertIsInstance(result, ExportArtifact)

    def test_custom_filename_option(self):
        result = self.adapter.export([INVOICE_1], options={"filename": "batch_2026_08.xlsx"})
        self.assertEqual(result.name, "batch_2026_08.xlsx")

    def test_custom_column_labels_override(self):
        custom_labels = {
            "invoice_number": "Số chứng từ",
            "supplier_name": "Tên đối tượng",
            "total_amount": "Tổng tiền",
        }
        result = self.adapter.export([INVOICE_1], options={"column_labels": custom_labels})
        wb = _load_workbook_from_artifact(result)
        ws = wb["Invoices"]
        header = [cell.value for cell in ws[1]]
        self.assertIn("Số chứng từ", header)
        self.assertIn("Tên đối tượng", header)
        self.assertIn("Tổng tiền", header)
        self.assertNotIn("Invoice Number", header)


if __name__ == "__main__":
    unittest.main()

