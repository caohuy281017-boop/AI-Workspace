"""XLSX Export Adapter — implements ExportProvider port.

Converts a list of approved invoice record dicts into a two-sheet Excel
workbook (.xlsx):
  - Sheet "Invoices"   : one row per invoice, scalar fields as columns.
  - Sheet "Line Items" : one row per line item, linked by invoice number.

Dependencies:
  openpyxl (MIT licence — recorded in docs/REPO_MAP.md)

GitNexus boundary: openpyxl is imported only here, inside adapters/.
The ExportArtifact returned is a pure domain object (bytes + metadata).
"""

from __future__ import annotations

import hashlib
import io
import logging
from datetime import datetime
from typing import Any, Mapping, Sequence

from core_shared.domain import ExportArtifact

logger = logging.getLogger(__name__)

_EXPORTER_NAME = "XLSXExportAdapter/openpyxl"

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

INVOICE_COLUMNS: list[tuple[str, str]] = [
    # (header label, dict key)
    ("Source File ID",   "source_file_id"),
    ("Invoice Number",   "invoice_number"),
    ("Invoice Date",     "invoice_date"),
    ("Supplier Name",    "supplier_name"),
    ("Supplier Tax ID",  "supplier_tax_id"),
    ("Currency",         "currency"),
    ("Subtotal",         "subtotal"),
    ("Tax Amount",       "tax_amount"),
    ("Total Amount",     "total_amount"),
]

LINE_ITEM_COLUMNS: list[tuple[str, str]] = [
    # (header label, dict key)
    ("Invoice Number", "invoice_number"),
    ("Description",    "description"),
    ("Quantity",       "quantity"),
    ("Unit Price",     "unit_price"),
    ("Amount",         "amount"),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _default_filename() -> str:
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    return f"invoices_{stamp}.xlsx"


def _style_header_row(ws: Any, num_cols: int) -> None:
    """Apply bold + light-grey fill to the header row."""
    try:
        from openpyxl.styles import Font, PatternFill  # type: ignore[import]

        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        bold_font = Font(bold=True)
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = bold_font
            cell.fill = header_fill
    except Exception:  # noqa: BLE001
        pass  # Styling is cosmetic — never block the export


def _auto_column_widths(ws: Any) -> None:
    """Set each column width based on the longest cell value."""
    try:
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
    except Exception:  # noqa: BLE001
        pass


# ---------------------------------------------------------------------------
# Public adapter
# ---------------------------------------------------------------------------

class XLSXExportAdapter:
    """``ExportProvider`` implementation that writes a two-sheet XLSX workbook.

    Usage::

        exporter = XLSXExportAdapter()
        artifact = exporter.export(approved_records, options={"filename": "batch.xlsx"})
        # artifact.content is the raw bytes of the .xlsx file
    """

    def export(
        self,
        records: Sequence[Mapping[str, Any]],
        *,
        options: Mapping[str, Any] | None = None,
    ) -> ExportArtifact:
        """Build and return the XLSX export artifact.

        Args:
            records: Sequence of approved invoice dicts (from ExtractionResult.values).
            options: Optional dict.  Supported keys:
                ``filename`` (str): override the default filename.

        Returns:
            ExportArtifact with media_type, name, content (bytes), and sha256.
        """
        try:
            import openpyxl  # type: ignore[import]
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for XLSX export. "
                "Install it with: pip install openpyxl"
            ) from exc

        opts = options or {}
        filename = opts.get("filename") or _default_filename()

        wb = openpyxl.Workbook()

        # ----------------------------------------------------------------
        # Sheet 1: Invoices
        # ----------------------------------------------------------------
        ws_inv = wb.active
        ws_inv.title = "Invoices"

        inv_headers = [col[0] for col in INVOICE_COLUMNS]
        ws_inv.append(inv_headers)
        _style_header_row(ws_inv, len(inv_headers))

        for record in records:
            row = [record.get(col[1]) for col in INVOICE_COLUMNS]
            ws_inv.append(row)

        _auto_column_widths(ws_inv)

        # ----------------------------------------------------------------
        # Sheet 2: Line Items
        # ----------------------------------------------------------------
        ws_items = wb.create_sheet("Line Items")

        item_headers = [col[0] for col in LINE_ITEM_COLUMNS]
        ws_items.append(item_headers)
        _style_header_row(ws_items, len(item_headers))

        for record in records:
            invoice_number = record.get("invoice_number")
            items = record.get("items") or []
            for item in items:
                item_dict = dict(item) if not isinstance(item, dict) else item
                row = [
                    invoice_number if col[1] == "invoice_number" else item_dict.get(col[1])
                    for col in LINE_ITEM_COLUMNS
                ]
                ws_items.append(row)

        _auto_column_widths(ws_items)

        # ----------------------------------------------------------------
        # Serialize to bytes
        # ----------------------------------------------------------------
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()
        sha256 = hashlib.sha256(content).hexdigest()

        logger.info(
            "XLSX export: %d invoices, filename=%s, size=%d bytes",
            len(records),
            filename,
            len(content),
        )

        return ExportArtifact(
            name=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=content,
            exporter=_EXPORTER_NAME,
            sha256=sha256,
        )
